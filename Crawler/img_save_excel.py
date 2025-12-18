# -*- coding: utf-8 -*-
import os
import re
import sys
import logging
import io
import olefile
import xlrd
import pandas as pd
from PIL import Image
from openpyxl import load_workbook, __version__ as openpyxl_version
import datetime
from path_utils import PathManager, setup_path_logging, get_path_manager

# 配置全局日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("excel_image_extractor.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)

# 可配置常量
FILE_NAME_COL = 5      # 文件名来源列号（Excel列号，从1开始）
FOLDER_NAME_COL = 9    # 文件名称列号（Excel列号，从1开始）
TARGET_COLS = [18, 19, 20, 21]  # 目标列号（Excel列号，从1开始）
MIN_IMAGE_SIZE = 1024  # 最小图片大小（字节）

def validate_environment():
    """验证运行环境"""
    logging.info(f"Python版本: {sys.version}")
    logging.info(f"openpyxl版本: {openpyxl_version}")


def process_embedded_images(excel_path, output_root, times=None):
    """主处理函数"""
    try:
        if not os.path.isfile(excel_path):
            raise FileNotFoundError(f"文件不存在: {os.path.abspath(excel_path)}")

        if excel_path.lower().endswith('.xlsx'):
            process_xlsx(excel_path, output_root, times)
        elif excel_path.lower().endswith('.xls'):
            process_xls(excel_path, output_root, times)
        else:
            raise ValueError("不支持的文件格式")

    except Exception as e:
        logging.error("处理过程发生异常", exc_info=True)
        raise


def process_xlsx(excel_path, output_root, times=None):
    """处理xlsx文件"""
    wb = load_workbook(excel_path, keep_vba=False)
    ws = wb.active

    # 读取Excel数据以获取主运营和SBS账单编号信息
    excel_data = None
    try:
        excel_data = pd.read_excel(excel_path)
        logging.info(f"成功读取Excel数据，共{len(excel_data)}行")
    except Exception as e:
        logging.warning(f"读取Excel数据失败，将使用默认路径结构: {e}")

    # 表头处理
    header_map = get_headers(ws, TARGET_COLS)

    # 图片映射
    image_map = {}
    images = ws.images if hasattr(ws, 'images') else getattr(ws, '_images', [])
    for img in images:
        if hasattr(img.anchor, '_from'):
            row = img.anchor._from.row + 1
            col = img.anchor._from.col + 1
            if col in TARGET_COLS:
                image_map[(row, col)] = img

    # 遍历行处理
    for row_idx in range(2, ws.max_row + 1):
        process_row(ws, row_idx, image_map, header_map, output_root, 'xlsx', excel_data, times)

    wb.close()


def process_xls(excel_path, output_root, times=None):
    """处理xls文件"""
    # 读取单元格数据
    wb = xlrd.open_workbook(excel_path)
    ws = wb.sheet_by_index(0)

    # 读取Excel数据以获取主运营和SBS账单编号信息
    excel_data = None
    try:
        excel_data = pd.read_excel(excel_path)
        logging.info(f"成功读取Excel数据，共{len(excel_data)}行")
    except Exception as e:
        logging.warning(f"读取Excel数据失败，将使用默认路径结构: {e}")

    # 动态获取有效列
    valid_cols = [col for col in TARGET_COLS if (col - 1) < ws.ncols]
    logging.info(f"有效目标列: {valid_cols}")

    # 获取表头
    header_map = get_headers(ws, valid_cols)

    # 提取图片数据
    images_data = extract_ole_images(excel_path)

    # 计算每行图片数
    images_per_row = len(valid_cols)
    if images_per_row == 0:
        raise ValueError("没有有效目标列可供处理")

    # 遍历行处理
    for row_idx in range(1, ws.nrows):
        process_xls_row(ws, row_idx, images_data, valid_cols, header_map, output_root, images_per_row, excel_data, times)


def process_row(ws, row_idx, image_map, header_map, output_root, file_type, excel_data=None, times=None):
    """处理单行数据，支持新的路径结构"""
    try:
        # 获取文件夹名称
        folder_cell = ws.cell(row=row_idx, column=FOLDER_NAME_COL) if file_type == 'xlsx' else ws.cell_value(row_idx, FOLDER_NAME_COL - 1)
        folder_name = str(folder_cell.value if file_type == 'xlsx' else folder_cell).strip() if folder_cell else ""
        
        if not folder_name or folder_name == "None":
            logging.warning(f"第{row_idx}行文件夹名称为空，跳过处理")
            return

        # 清理文件夹名称
        clean_folder = clean_folder_name(folder_name, row_idx)
        
        # 尝试使用新的路径结构
        target_dir = output_root
        if excel_data is not None and times is not None:
            try:
                # 获取当前行的主运营和SBS账单编号信息
                # row_idx是Excel中的行号（从2开始，因为第1行是表头）
                # 转换为pandas DataFrame的索引（从0开始）
                pandas_row_idx = row_idx - 2  # Excel第2行对应pandas的第0行
                if pandas_row_idx >= 0 and pandas_row_idx < len(excel_data):
                    row_data = excel_data.iloc[pandas_row_idx]
                    
                    # 智能识别主运营列（可能是'主运营'或第一列）
                    operator_col = None
                    sbs_col = None
                    
                    # 查找主运营列
                    if '主运营' in row_data:
                        operator_col = '主运营'
                    elif len(row_data) > 0:
                        # 使用第一列作为主运营列
                        operator_col = row_data.index[0]
                    
                    # 查找SBS账单编号列
                    if 'SBS账单编号' in row_data:
                        sbs_col = 'SBS账单编号'
                    
                    if operator_col is not None and sbs_col is not None:
                        operator = str(row_data[operator_col]).strip()
                        sbs_number = str(row_data[sbs_col]).strip()
                        
                        if operator and operator != 'nan' and sbs_number and sbs_number != 'nan':
                            path_manager = get_path_manager(times)
                            sbs_base_path = path_manager.get_operator_sbs_path(operator, sbs_number)
                            target_dir = os.path.join(sbs_base_path, clean_folder)
                            logging.info(f"使用新路径结构: {target_dir}")
                        else:
                            target_dir = os.path.join(output_root, clean_folder)
                            logging.info(f"主运营或SBS账单编号为空，使用默认路径: {target_dir}")
                    else:
                        target_dir = os.path.join(output_root, clean_folder)
                        logging.info(f"缺少必要列，使用默认路径: {target_dir}")
                else:
                    target_dir = os.path.join(output_root, clean_folder)
                    logging.warning(f"行索引超出范围，使用默认路径: {target_dir}")
                    
            except Exception as e:
                logging.warning(f"获取新路径结构失败，使用默认路径: {e}")
                target_dir = os.path.join(output_root, clean_folder)
        else:
            target_dir = os.path.join(output_root, clean_folder)
        
        # 确保目录存在
        os.makedirs(target_dir, exist_ok=True)
        logging.info(f"第{row_idx}行图片保存到: {target_dir}")
        
        # 处理目标列的图片
        for col in TARGET_COLS:
            if (row_idx, col) in image_map:
                img = image_map[(row_idx, col)]
                header = header_map.get(col, f"列{col}")
                save_image(img, target_dir, row_idx, col, header, file_type)
                
    except Exception as e:
        logging.error(f"处理第{row_idx}行时发生错误: {e}", exc_info=True)

def process_xls_row(ws, row_idx, images_data, valid_cols, header_map, output_root, images_per_row, excel_data=None, times=None):
    """xls专用行处理，支持新的路径结构"""
    try:
        # 获取文件夹名称
        folder_name = clean_folder_name(ws.cell(row_idx, FOLDER_NAME_COL - 1).value, row_idx + 1)
        
        if not folder_name or folder_name == "None":
            logging.warning(f"第{row_idx + 1}行文件夹名称为空，跳过处理")
            return

        # 尝试使用新的路径结构
        target_dir = output_root
        if excel_data is not None and times is not None:
            try:
                # 获取当前行的主运营和SBS账单编号信息
                # row_idx是xlrd中的行号（从1开始，因为第0行是表头）
                # 转换为pandas DataFrame的索引（从0开始）
                pandas_row_idx = row_idx - 1  # xlrd第1行对应pandas的第0行
                if pandas_row_idx >= 0 and pandas_row_idx < len(excel_data):
                    row_data = excel_data.iloc[pandas_row_idx]
                    
                    # 智能识别主运营列（可能是'主运营'或第一列）
                    operator_col = None
                    sbs_col = None
                    
                    # 查找主运营列
                    if '主运营' in row_data:
                        operator_col = '主运营'
                    elif len(row_data) > 0:
                        # 使用第一列作为主运营列
                        operator_col = row_data.index[0]
                    
                    # 查找SBS账单编号列
                    if 'SBS账单编号' in row_data:
                        sbs_col = 'SBS账单编号'
                    
                    if operator_col is not None and sbs_col is not None:
                        operator = str(row_data[operator_col]).strip()
                        sbs_number = str(row_data[sbs_col]).strip()
                        
                        if operator and operator != 'nan' and sbs_number and sbs_number != 'nan':
                            path_manager = get_path_manager(times)
                            sbs_base_path = path_manager.get_operator_sbs_path(operator, sbs_number)
                            target_dir = os.path.join(sbs_base_path, folder_name)
                            logging.info(f"使用新路径结构: {target_dir}")
                        else:
                            target_dir = os.path.join(output_root, folder_name)
                            logging.info(f"主运营或SBS账单编号为空，使用默认路径: {target_dir}")
                    else:
                        target_dir = os.path.join(output_root, folder_name)
                        logging.info(f"缺少必要列，使用默认路径: {target_dir}")
                else:
                    target_dir = os.path.join(output_root, folder_name)
                    logging.warning(f"行索引超出范围，使用默认路径: {target_dir}")
                    
            except Exception as e:
                logging.warning(f"获取新路径结构失败，使用默认路径: {e}")
                target_dir = os.path.join(output_root, folder_name)
        else:
            target_dir = os.path.join(output_root, folder_name)

        # 确保目录存在
        os.makedirs(target_dir, exist_ok=True)
        logging.info(f"第{row_idx + 1}行图片保存到: {target_dir}")

        start_idx = (row_idx - 1) * images_per_row
        for idx_offset, col in enumerate(valid_cols):
            img_idx = start_idx + idx_offset
            if img_idx >= len(images_data):
                break

            header = header_map.get(col, f"Col{col}")
            save_xls_image(
                images_data[img_idx],
                target_dir,
                row_idx + 1,
                col,
                header
            )
            
    except Exception as e:
        logging.error(f"处理第{row_idx + 1}行时发生错误: {e}", exc_info=True)


def extract_ole_images(excel_path):
    """从OLE结构提取图片"""
    ole = olefile.OleFileIO(excel_path)
    images_data = []

    # 扩展识别条件
    stream_patterns = ['pictures', 'mbd', 'image', 'blip']
    for stream in ole.listdir():
        stream_name = stream[0].lower()
        if any(p in stream_name for p in stream_patterns):
            data = ole.openstream(stream).read()
            if len(data) >= MIN_IMAGE_SIZE:
                images_data.append(data)
                logging.debug(f"发现图片流: {stream} ({len(data) // 1024}KB)")

    ole.close()
    logging.info(f"共提取到{len(images_data)}张候选图片")
    return images_data


def get_headers(ws, target_cols):
    """通用表头获取"""
    header_map = {}
    for col in target_cols:
        try:
            if isinstance(ws, xlrd.sheet.Sheet):  # xlrd处理
                cell = ws.cell(0, col - 1)
            else:  # openpyxl处理
                cell = ws.cell(row=1, column=col)
            header_map[col] = clean_header_name(cell.value)
        except Exception as e:
            logging.warning(f"获取列{col}表头失败: {str(e)}")
            header_map[col] = f"Col{col}"
    return header_map


def safe_filename(name):
    """文件名安全处理（全平台兼容）"""
    if not name:
        return "unnamed"

    # 替换危险字符
    name = str(name).strip()
    name = re.sub(r'[\\/*?:"<>|]', '', name)
    name = re.sub(r'[\x00-\x1F\x7F]', '', name)  # 去除控制字符

    # 处理特殊符号
    name = name.replace(' ', '_')
    name = name.replace('.', '_')

    # 限制长度
    return name[:40]

def save_image(img, target_dir, row, col, header, file_type):
    """增强保存函数"""
    try:
        base_name = safe_filename(os.path.basename(target_dir))
        header_clean = safe_filename(header)
        filename = f"{base_name} {header_clean}.png"  # 添加行列标识

        # 处理长路径
        output_path = os.path.join(target_dir, filename)
        if os.name == 'nt' and len(output_path) > 240:
            output_path = '\\\\?\\' + os.path.abspath(output_path)

        # 确保父目录存在（二次验证）
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # 写入文件
        with open(output_path, "wb") as f:
            data = img._data() if file_type == 'xlsx' else img
            f.write(data)

        logging.info(f"行{row}列{col} 保存成功 → {filename}")
    except Exception as e:
        logging.error(f"行{row}列{col} 保存失败 | 路径: {output_path} | 错误: {str(e)}")


def save_xls_image(img_data, target_dir, row, col, header):
    """xls专用图片保存"""
    base_name = f"{os.path.basename(target_dir)}_{header}"
    try:
        img = Image.open(io.BytesIO(img_data))
        img.save(os.path.join(target_dir, f"{base_name}.png"))
        logging.info(f"行{row}列{col} → 保存为PNG")
    except Exception as e:
        logging.warning(f"行{row}列{col} PNG解析失败: {str(e)[:50]}...")
        with open(os.path.join(target_dir, f"{base_name}.emf"), "wb") as f:
            f.write(img_data)
        logging.info(f"行{row}列{col} → 保存为EMF")


def clean_folder_name(raw_name, row_number):
    """清理文件夹名称"""
    if not raw_name:
        return f"未命名_{row_number}"
    cleaned = re.sub(r'[\\/*?:"<>|]', '', str(raw_name)).strip()[:50]
    return cleaned or f"未命名_{row_number}"


def clean_header_name(raw_header):
    """清洗表头名称"""
    if not raw_header:
        return "未命名表头"
    cleaned = re.sub(r'[\\/*?:"<>|]', '', str(raw_header))
    cleaned = re.sub(r'_+.*', '', cleaned)
    return cleaned.strip()[:20] or "未命名表头"


if __name__ == "__main__":
    validate_environment()
    times = input("请输入本次申诉时间：")
    if times == "":
        times = datetime.datetime.now().strftime("%Y.%m.%d")
    
    # 设置路径日志
    setup_path_logging(times)
    
    # 使用路径管理器
    path_manager = get_path_manager(times)
    EXCEL_PATH = path_manager.get_material_file_path()
    OUTPUT_DIR = path_manager.base_dir
    
    logging.info(f"开始处理Excel图片提取，时间: {times}")
    logging.info(f"Excel文件路径: {EXCEL_PATH}")
    logging.info(f"输出目录: {OUTPUT_DIR}")
    
    try:
        # 验证路径
        if not path_manager.validate_path(os.path.dirname(EXCEL_PATH)):
            raise ValueError(f"Excel文件目录无效: {os.path.dirname(EXCEL_PATH)}")
        
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        
        # 处理图片提取，传递times参数以支持新路径结构
        process_embedded_images(EXCEL_PATH, OUTPUT_DIR, times)
        
        logging.info("处理完成！3秒后自动打开结果目录...")
        print(f"图片提取完成！文件保存在: {OUTPUT_DIR}")
        
        # 记录路径处理日志
        log_file = path_manager.get_log_file_path("img_save_excel")
        logging.info(f"详细日志保存在: {log_file}")
        
        os.startfile(OUTPUT_DIR)
    except Exception as e:
        logging.critical(f"致命错误: {str(e)}")
        print(f"处理失败: {str(e)}")
    finally:
        input("按回车键退出程序...")